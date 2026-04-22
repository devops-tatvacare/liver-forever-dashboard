#!/usr/bin/env python3
"""
Build src/data/doctors.json from the Zydus master Excel's "Doctor Reference" sheet.

Only keeps doctor IDs that appear in src/data/scans.json so the bundle stays small.

Usage:
  python3 scripts/build-doctors.py <path-to-xlsx> <output-json>

Example:
  python3 scripts/build-doctors.py \
    "~/OneDrive/.../Liver Forever - FibroScan Uploads Data Apr 17.xlsx" \
    src/data/doctors.json

Requires: `pip install openpyxl`.
"""
import json
import pathlib
import sys

import openpyxl  # type: ignore


def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__)
        return 2
    xlsx_path = pathlib.Path(sys.argv[1]).expanduser()
    out_path = pathlib.Path(sys.argv[2])

    root = pathlib.Path(__file__).resolve().parent.parent
    scans_path = root / "src" / "data" / "scans.json"
    if not scans_path.exists():
        print(f"!! {scans_path} not found — run the MySQL export first.")
        return 1

    scans = json.loads(scans_path.read_text())
    used = {s["doctor_id"] for s in scans if s.get("doctor_id")}
    print(f"Distinct doctor IDs in scans.json: {len(used)}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Doctor Reference"]
    header = [c.value for c in ws[1]]
    idx = {n: header.index(n) for n in
           ["drCode", "doctorName", "clinicName", "cityName", "stateName", "employeeName", "hqDesc"]}

    out: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[idx["drCode"]]
        if dr is None:
            continue
        key = str(dr).strip()
        if key not in used or key in out:
            continue
        def cell(name: str) -> str:
            v = row[idx[name]]
            return str(v).strip().title() if v else ""
        out[key] = {
            "name":   cell("doctorName"),
            "clinic": cell("clinicName"),
            "city":   cell("cityName"),
            "state":  cell("stateName"),
            "mr":     cell("employeeName"),
            "hq":     cell("hqDesc"),
        }

    matched = len(out)
    print(f"Matched {matched} / {len(used)} doctor IDs")
    out_path.write_text(json.dumps(out, ensure_ascii=False))
    print(f"Wrote {out_path} ({out_path.stat().st_size // 1024} KB)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
