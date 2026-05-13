#!/usr/bin/env python3
"""
Build the Liver Forever Registration Funnel aggregates workbook.

Inputs:
  - src/data/registration-funnel.json
  - <repo>/.playwright-mcp/registration-funnel-full.png

Outputs (under <repo>/reports/):
  - Liver_Forever_Registration_Funnel_<YYYY-MM-DD>.xlsx

Tabs:
  - Overview      — KPIs + narrative + all aggregate tables (no patient-level data)
  - Funnel Snapshot — embedded screenshot of the dashboard page

Style tokens match the Fibroscan Outcomes report for visual coherence.
"""
from __future__ import annotations

import json
import pathlib
from datetime import date

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


ROOT = pathlib.Path(__file__).resolve().parent.parent
REPO = ROOT.parent
DATA = ROOT / "src" / "data" / "registration-funnel.json"
SHOT = REPO / ".playwright-mcp" / "registration-funnel-full.png"
OUT_DIR = REPO / "reports"
OUT_PATH = OUT_DIR / f"Liver_Forever_Registration_Funnel_{date.today():%Y-%m-%d}.xlsx"


# Design tokens (same as Fibroscan report)
BRAND        = "7030A0"
BRAND_DEEP   = "2D1B69"
BRAND_SOFT   = "F2EDF8"
ACCENT_INDIGO= "6366F1"
ACCENT_CYAN  = "06B6D4"
INK          = "252530"
INK_MUTED    = "6B6B7B"
GREY_HEADER  = "F5F4F9"
GREY_ROW     = "FBFAFF"
BORDER       = "E8E7EE"
SUCCESS      = "16A34A"
SUCCESS_LITE = "D1FAE5"
ERROR        = "DC2626"
ERROR_LITE   = "FEE2E2"
WARNING      = "CA8A04"
WARNING_LITE = "FEF3C7"

FONT_H1      = Font(name="Calibri", size=22, bold=True, color=BRAND_DEEP)
FONT_H2      = Font(name="Calibri", size=14, bold=True, color=INK)
FONT_EYEBROW = Font(name="Calibri", size=9, bold=True, color=BRAND)
FONT_BODY    = Font(name="Calibri", size=11, color=INK)
FONT_MUTED   = Font(name="Calibri", size=10, color=INK_MUTED)
FONT_TBL_HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_KPI_LBL = Font(name="Calibri", size=9, bold=True, color=INK_MUTED)

FILL_TBL_HEADER = PatternFill("solid", start_color=BRAND_DEEP)
FILL_KPI        = PatternFill("solid", start_color=GREY_HEADER)
FILL_WARN       = PatternFill("solid", start_color=WARNING_LITE)
FILL_SUCCESS    = PatternFill("solid", start_color=SUCCESS_LITE)

THIN = Side(border_style="thin", color=BORDER)
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def ws_setup(ws: Worksheet, tab_color: str = BRAND) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color


def set_col_widths(ws: Worksheet, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_header_row(ws: Worksheet, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = FONT_TBL_HDR
        c.fill = FILL_TBL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 28


def stripe_rows(ws: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    fill = PatternFill("solid", start_color=GREY_ROW)
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 0:
            for col in range(start_col, end_col + 1):
                ws.cell(row=r, column=col).fill = fill


def kpi_block(ws: Worksheet, row: int, col: int, label: str, value, fmt: str | None = None, color: str | None = None) -> None:
    lbl = ws.cell(row=row, column=col, value=label)
    lbl.font = FONT_KPI_LBL
    lbl.fill = FILL_KPI
    lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lbl.border = BOX

    v = ws.cell(row=row + 1, column=col, value=value)
    v.font = Font(name="Calibri", size=20, bold=True, color=color or INK)
    v.fill = FILL_KPI
    v.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    v.border = BOX
    if fmt:
        v.number_format = fmt

    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 30


def banner(ws: Worksheet, row: int, eyebrow: str, title: str, subtitle: str = "") -> int:
    ws.cell(row=row, column=2, value=eyebrow).font = FONT_EYEBROW
    ws.cell(row=row + 1, column=2, value=title).font = FONT_H1
    if subtitle:
        c = ws.cell(row=row + 2, column=2, value=subtitle)
        c.font = FONT_MUTED
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row + 2, start_column=2, end_row=row + 2, end_column=7)
        return row + 4
    return row + 3


def section_title(ws: Worksheet, row: int, col: int, text: str) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.font = FONT_H2


def write_table(ws: Worksheet, start_row: int, headers: list[str], rows: list[list], col_count: int) -> int:
    """Write a styled table. Returns row index after table (header + body + 1)."""
    hdr_row = start_row
    for i, h in enumerate(headers):
        ws.cell(row=hdr_row, column=2 + i, value=h)
    style_header_row(ws, hdr_row, 2, 2 + col_count - 1)

    for ri, row in enumerate(rows):
        r = hdr_row + 1 + ri
        for ci, val in enumerate(row):
            cell = ws.cell(row=r, column=2 + ci, value=val)
            cell.border = BOX
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.alignment = Alignment(horizontal="right")
                if ci == 0:
                    cell.alignment = Alignment(horizontal="left")
    stripe_rows(ws, hdr_row + 1, hdr_row + len(rows), 2, 2 + col_count - 1)
    return hdr_row + len(rows) + 2


def pct(n: int, d: int) -> str:
    return f"{(100*n/d):.1f}%" if d else "—"


def build_overview(wb: Workbook, data: dict) -> None:
    summary = data["summary"]
    month_funnel = data["month_funnel"]
    cohorts = data["cohorts"]
    sources = data["by_source"]
    notes = data["notes"]
    src_file = data.get("source_file", "")
    gen_at = data.get("generated_at", "")

    default = wb.active
    if default is not None:
        wb.remove(default)
    ws = wb.create_sheet("Overview")
    ws_setup(ws)
    set_col_widths(ws, [2, 28, 18, 16, 16, 16, 16, 16, 4])

    row = 2
    row = banner(
        ws,
        row,
        eyebrow="LIVER FOREVER · REGISTRATION FUNNEL",
        title="From Sign-up to Verified Eligibility",
        subtitle=(
            f"Prepared {date.today():%d %b %Y}. Source: {src_file}. Aggregates only — no patient-level identifiers in this workbook."
        ),
    )

    # KPI grid (4 cols)
    section_title(ws, row, 2, "Headline funnel numbers")
    row += 2
    kpi_block(ws, row, 2, "SIGNED UP",          summary["signed_up"],      fmt="#,##0")
    kpi_block(ws, row, 3, "APP LOGIN (≥1)",     summary["app_login"],      fmt="#,##0")
    kpi_block(ws, row, 4, "UPLOADED ≥1 DOC",    summary["any_doc"],        fmt="#,##0")
    kpi_block(ws, row, 5, "BOTH DOCS UPLOADED", summary["both_uploaded"],  fmt="#,##0")
    kpi_block(ws, row, 6, "BOTH DOCS VERIFIED", summary["both_approved"],  fmt="#,##0", color=SUCCESS)
    kpi_block(ws, row, 7, "PENDING REVIEW",     summary["pending_verification"], fmt="#,##0", color=WARNING)
    row += 3

    # Narrative
    section_title(ws, row, 2, "The story so far")
    row += 1
    bot = next((s for s in sources if s["source"] == "bot"), None)
    web = next((s for s in sources if s["source"] == "web"), None)
    bot_rate = pct(bot["both_approved"], bot["sign_ups"]) if bot else "—"
    web_rate = pct(web["both_approved"], web["sign_ups"]) if web else "—"
    story_lines = [
        (
            f"Of {summary['signed_up']:,} sign-ups, only {summary['both_approved']:,} "
            f"({pct(summary['both_approved'], summary['signed_up'])}) have cleared the verification gate that unlocks coaching and refills."
        ),
        (
            f"{(100 - 100*summary['app_login']/summary['signed_up']):.0f}% of sign-ups never log into the app; only "
            f"{summary['any_doc']:,} ({pct(summary['any_doc'], summary['signed_up'])}) ever upload a single document."
        ),
        (
            f"{summary['pending_verification']:,} patients are sitting fully submitted but with neither document approved yet — "
            f"{summary['docs_under_review_total']:,} documents total are in 'Under Review' status."
        ),
    ]
    if bot and web:
        story_lines.append(
            f"Web sign-ups convert to verification at {web_rate} vs WhatsApp bot at {bot_rate} — the bot is the volume firehose, the web form is where intent is highest."
        )
    story_lines.append(
        f"The program tags {notes['coach_eligible_tag_total']} patients 'All Coaches Eligible' but only "
        f"{notes['coach_eligible_with_both_approved']} of those have both Month-1 docs verified — reconcile the operational definition before reporting eligibility externally."
    )
    for line in story_lines:
        ws.cell(row=row, column=2, value="• " + line).font = FONT_BODY
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 32
        row += 1
    row += 1

    # Funnel A
    section_title(ws, row, 2, "Funnel A — Overall patient journey")
    row += 2
    stages = [
        ("Signed Up",         summary["signed_up"],     summary["signed_up"]),
        ("App Login",         summary["app_login"],     summary["signed_up"]),
        ("≥1 Doc Uploaded",   summary["any_doc"],       summary["app_login"]),
        ("Both Docs Uploaded",summary["both_uploaded"], summary["any_doc"]),
        ("Both Docs Verified",summary["both_approved"], summary["both_uploaded"]),
    ]
    top = summary["signed_up"]
    hdr_row = row
    headers = ["Stage", "Patients", "% of Sign-ups", "% of Previous Step"]
    for i, h in enumerate(headers):
        ws.cell(row=hdr_row, column=2 + i, value=h)
    style_header_row(ws, hdr_row, 2, 5)
    for i, (label, value, prev) in enumerate(stages):
        r = hdr_row + 1 + i
        ws.cell(row=r, column=2, value=label).font = FONT_BODY
        ws.cell(row=r, column=3, value=value).number_format = "#,##0"
        ws.cell(row=r, column=4, value=value / top if top else 0).number_format = "0.0%"
        ws.cell(row=r, column=5, value=(value / prev) if prev and i > 0 else 1).number_format = "0.0%"
        for col in range(2, 6):
            ws.cell(row=r, column=col).border = BOX
    stripe_rows(ws, hdr_row + 1, hdr_row + len(stages), 2, 5)
    # Highlight final row
    for col in range(2, 6):
        ws.cell(row=hdr_row + len(stages), column=col).fill = FILL_SUCCESS
    row = hdr_row + len(stages) + 2

    # Funnel B — month progression
    section_title(ws, row, 2, "Funnel B — Program stage progression (Month 1 → Month 6)")
    row += 2
    hdr_row = row
    headers = ["Stage", "Uploaded (both)", "Verified (both)", "Verify rate"]
    for i, h in enumerate(headers):
        ws.cell(row=hdr_row, column=2 + i, value=h)
    style_header_row(ws, hdr_row, 2, 5)
    for i, m in enumerate(month_funnel):
        r = hdr_row + 1 + i
        ws.cell(row=r, column=2, value=f"Month {m['month']}").font = FONT_BODY
        ws.cell(row=r, column=3, value=m["uploaded"]).number_format = "#,##0"
        ws.cell(row=r, column=4, value=m["verified"]).number_format = "#,##0"
        if m["uploaded"]:
            ws.cell(row=r, column=5, value=m["verified"] / m["uploaded"]).number_format = "0.0%"
        else:
            ws.cell(row=r, column=5, value="—").alignment = Alignment(horizontal="right")
        for col in range(2, 6):
            ws.cell(row=r, column=col).border = BOX
    stripe_rows(ws, hdr_row + 1, hdr_row + len(month_funnel), 2, 5)
    row = hdr_row + len(month_funnel) + 2

    # Stage-tag mismatch callout
    callout = (
        f"⚠ Stage-tag vs. verification mismatch — {notes['coach_eligible_tag_total']} patients tagged 'All Coaches Eligible'; "
        f"only {notes['coach_eligible_with_both_approved']} have both Month-1 docs verified."
    )
    c = ws.cell(row=row, column=2, value=callout)
    c.font = Font(name="Calibri", size=10, bold=True, color=INK)
    c.fill = FILL_WARN
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    ws.row_dimensions[row].height = 26
    row += 2

    # Source breakdown
    section_title(ws, row, 2, "Acquisition channel breakdown")
    row += 2
    hdr_row = row
    headers = ["Source", "Sign-ups", "App Login", "≥1 Doc", "Both Uploaded", "Both Verified", "Verify Rate"]
    for i, h in enumerate(headers):
        ws.cell(row=hdr_row, column=2 + i, value=h)
    style_header_row(ws, hdr_row, 2, 8)
    source_labels = {"bot": "WhatsApp Bot", "web": "Web Form", "app": "App"}
    for i, s in enumerate(sources):
        r = hdr_row + 1 + i
        ws.cell(row=r, column=2, value=source_labels.get(s["source"], s["source"])).font = FONT_BODY
        ws.cell(row=r, column=3, value=s["sign_ups"]).number_format = "#,##0"
        ws.cell(row=r, column=4, value=s["app_login"]).number_format = "#,##0"
        ws.cell(row=r, column=5, value=s["any_doc"]).number_format = "#,##0"
        ws.cell(row=r, column=6, value=s["both_uploaded"]).number_format = "#,##0"
        ws.cell(row=r, column=7, value=s["both_approved"]).number_format = "#,##0"
        ws.cell(row=r, column=8, value=(s["both_approved"] / s["sign_ups"]) if s["sign_ups"] else 0).number_format = "0.0%"
        for col in range(2, 9):
            ws.cell(row=r, column=col).border = BOX
    stripe_rows(ws, hdr_row + 1, hdr_row + len(sources), 2, 8)
    row = hdr_row + len(sources) + 2

    # Cohort table
    section_title(ws, row, 2, "Cohort movement — by sign-up month")
    row += 2
    hdr_row = row
    headers = ["Sign-up Month", "Sign-ups", "App Login", "≥1 Doc", "Both Uploaded", "Both Verified", "Verify Rate"]
    for i, h in enumerate(headers):
        ws.cell(row=hdr_row, column=2 + i, value=h)
    style_header_row(ws, hdr_row, 2, 8)
    totals = {"signups": 0, "logged_in": 0, "any_doc": 0, "both_uploaded": 0, "both_approved": 0}
    for c in cohorts:
        totals["signups"] += c["signups"]
        totals["logged_in"] += c["logged_in"]
        totals["any_doc"] += c["any_doc"]
        totals["both_uploaded"] += c["both_uploaded"]
        totals["both_approved"] += c["both_approved"]
    for i, c in enumerate(reversed(cohorts)):
        r = hdr_row + 1 + i
        ws.cell(row=r, column=2, value=c["signup_month"]).font = FONT_BODY
        ws.cell(row=r, column=3, value=c["signups"]).number_format = "#,##0"
        ws.cell(row=r, column=4, value=c["logged_in"]).number_format = "#,##0"
        ws.cell(row=r, column=5, value=c["any_doc"]).number_format = "#,##0"
        ws.cell(row=r, column=6, value=c["both_uploaded"]).number_format = "#,##0"
        ws.cell(row=r, column=7, value=c["both_approved"]).number_format = "#,##0"
        ws.cell(row=r, column=8, value=(c["both_approved"] / c["signups"]) if c["signups"] else 0).number_format = "0.0%"
        for col in range(2, 9):
            ws.cell(row=r, column=col).border = BOX
    stripe_rows(ws, hdr_row + 1, hdr_row + len(cohorts), 2, 8)
    # Total row
    tr = hdr_row + len(cohorts) + 1
    ws.cell(row=tr, column=2, value="Total").font = Font(bold=True, color=INK)
    ws.cell(row=tr, column=3, value=totals["signups"]).number_format = "#,##0"
    ws.cell(row=tr, column=4, value=totals["logged_in"]).number_format = "#,##0"
    ws.cell(row=tr, column=5, value=totals["any_doc"]).number_format = "#,##0"
    ws.cell(row=tr, column=6, value=totals["both_uploaded"]).number_format = "#,##0"
    ws.cell(row=tr, column=7, value=totals["both_approved"]).number_format = "#,##0"
    ws.cell(row=tr, column=8, value=(totals["both_approved"] / totals["signups"]) if totals["signups"] else 0).number_format = "0.0%"
    for col in range(2, 9):
        cell = ws.cell(row=tr, column=col)
        cell.font = Font(bold=True, color=INK)
        cell.fill = FILL_KPI
        cell.border = BOX
    row = tr + 3

    # Footer
    ws.cell(row=row, column=2, value=f"Generated {gen_at} · TatvaCare × Zydus Healthcare confidential.").font = FONT_MUTED


def build_snapshot(wb: Workbook, screenshot_path: pathlib.Path) -> None:
    ws = wb.create_sheet("Funnel Snapshot")
    ws_setup(ws, tab_color=ACCENT_INDIGO)
    set_col_widths(ws, [2.0] + [12.0] * 12)

    ws.cell(row=2, column=2, value="DASHBOARD SNAPSHOT").font = FONT_EYEBROW
    ws.cell(row=3, column=2, value="Liver Forever — Registration Funnel").font = FONT_H1
    ws.cell(row=4, column=2, value=f"Captured {date.today():%d %b %Y}. Live dashboard: http://localhost:5174/").font = FONT_MUTED

    if not screenshot_path.exists():
        ws.cell(row=6, column=2, value=f"Screenshot not found at {screenshot_path}").font = FONT_BODY
        return

    img = XLImage(str(screenshot_path))
    # Scale down to fit ~1100px wide for readability while keeping aspect ratio.
    target_width = 1100
    scale = target_width / img.width
    img.width = target_width
    img.height = int(img.height * scale)
    ws.add_image(img, "B6")


def main() -> int:
    data = json.loads(DATA.read_text())
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    build_overview(wb, data)
    build_snapshot(wb, SHOT)

    wb.save(OUT_PATH)
    print(f"✓ {OUT_PATH}  ({OUT_PATH.stat().st_size // 1024} KB)")
    print(f"  sheets: {wb.sheetnames}")
    print(f"  generated_at: {data.get('generated_at')}")
    print(f"  signed_up={data['summary']['signed_up']:,}  both_verified={data['summary']['both_approved']:,}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
